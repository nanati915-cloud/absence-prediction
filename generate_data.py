"""
機械学習モデル検証用のサンプル欠席データを生成する。

年齢・登園期間・季節性などの傾向を考慮した
疑似的な園児欠席データを作成し、
分析処理で利用するExcelファイルとして出力する。
"""

import numpy as np
import random
from openpyxl import Workbook

# =================================================
# ■ 1. 乱数シードの固定（データの再現性担保）
# =================================================
np.random.seed(20)
random.seed(20)

CHILDREN = ["子供001", "子供002"]

BASE_PROFILE = {
    "子供001": {"base": 4.0, "vol": 1.2, "season_sens": 1.0, "start_age_months": 11},
    "子供002": {"base": 3.0, "vol": 0.9, "season_sens": 0.85, "start_age_months": 24},
}

# =================================================
# ■ 2. ビジネスルールと確率ノイズ
# =================================================
def clamp(v, min_v=0, max_v=15):
    return max(min_v, min(max_v, int(round(v))))

def maybe_blank(p=0.05):
    return random.random() < p

def seasonal(month):
    if month in [1, 2]: return 1.8
    if month in [12]: return 1.4
    if month in [7, 8]: return 0.6
    return 1.0

def noise(scale):
    return np.random.normal(0, scale)

# =================================================
# ■ 3. 時系列マスターデータの生成
# =================================================
def generate_master_timeline(length=61):
    master_records = []
    start_year, start_month = 2021, 6

    for i in range(length):
        current_total_months = (start_year * 12 + (start_month - 1)) + i
        year = current_total_months // 12
        month = (current_total_months % 12) + 1
        
        record = {"index": i, "year_month_str": f"{year}年{month}月", "month": month, "data": {}}
        
        for child in CHILDREN:
            p = BASE_PROFILE[child]
            attendance_month = i
            current_age_months = p["start_age_months"] + i
            
            factor_age = max(0.4, 1.8 - current_age_months * 0.02)
            factor_seasonal = seasonal(month) * (1.0 - attendance_month * 0.003)
            factor = (factor_age + factor_seasonal) / 2.0
            val = (p["base"] * factor) + noise(p["vol"])
            final_val = clamp(val)
            
            if maybe_blank(0.05):
                final_val = None
                
            record["data"][child] = final_val
            
        master_records.append(record)
        
    return master_records

# =================================================
# ■ 4. マスターデータからの各軸への安全なマッピング・出力
# =================================================
def export_excel():
    master_data = generate_master_timeline(61)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "病欠集約"

    # 軸3. 月間（J列〜L列）
    ws["J2"], ws["K2"], ws["L2"] = "年月", "子供001", "子供002"
    for i, record in enumerate(master_data):
        ws.cell(row=3+i, column=10, value=record["year_month_str"])
        ws.cell(row=3+i, column=11, value=record["data"]["子供001"])
        ws.cell(row=3+i, column=12, value=record["data"]["子供002"])

    # 軸2. 登園月数（F列〜H列）
    ws["F2"], ws["G2"], ws["H2"] = "登園月数", "子供001", "子供002"
    for i, record in enumerate(master_data):
        ws.cell(row=3+i, column=6, value=f"{i // 12}年{i % 12}ヶ月")
        ws.cell(row=3+i, column=7, value=record["data"]["子供001"])
        ws.cell(row=3+i, column=8, value=record["data"]["子供002"])

    # 軸1. 年齢（B列〜D列）
    ws["B2"], ws["C2"], ws["D2"] = "年齢", "子供001", "子供002"
    p001 = BASE_PROFILE["子供001"]
    for i, record in enumerate(master_data):
        c1_age = p001["start_age_months"] + i
        ws.cell(row=3+i, column=2, value=f"{c1_age // 12}歳{c1_age % 12}ヶ月")
        ws.cell(row=3+i, column=3, value=record["data"]["子供001"])
        ws.cell(row=3+i, column=4, value=record["data"]["子供002"])

    # 軸4. 月別サマリー（N列〜P列）※合計値を出力
    ws["N2"], ws["O2"], ws["P2"] = "月", "子供001", "子供002"
    for m in range(1, 13):
        ws.cell(row=2+m, column=14, value=f"{m}月")
        for c_idx, child in enumerate(CHILDREN):
            vals = [r["data"][child] for r in master_data if r["month"] == m and r["data"][child] is not None]
            ws.cell(row=2+m, column=15+c_idx, value=sum(vals))

    # 軸5. 学年（R列〜T列）※合計値を出力
    ws["R2"], ws["S2"], ws["T2"] = "クラス", "子供001", "子供002"
    classes = ["0歳児クラス", "1歳児クラス", "2歳児クラス", "3歳児クラス", "4歳児クラス", "5歳児クラス", "小学1年生"]
    for cls_idx, cls_name in enumerate(classes):
        ws.cell(row=3+cls_idx, column=18, value=cls_name)
        for c_idx, child in enumerate(CHILDREN):
            vals = []
            for r in master_data:
                current_age = BASE_PROFILE[child]["start_age_months"] + r["index"]
                child_class_idx = (current_age - 4) // 12 if current_age >= 4 else 0
                if child_class_idx == cls_idx and r["data"][child] is not None:
                    vals.append(r["data"][child])
            ws.cell(row=3+cls_idx, column=19+c_idx, value=sum(vals))

    # 軸6. 年数（V列〜X列）※合計値を出力
    ws["V2"], ws["W2"], ws["X2"] = "年数", "子供001", "子供002"
    for y in range(6):
        ws.cell(row=3+y, column=22, value=f"{y}年")
        for c_idx, child in enumerate(CHILDREN):
            vals = [r["data"][child] for r in master_data if r["index"] // 12 == y and r["data"][child] is not None]
            ws.cell(row=3+y, column=23+c_idx, value=sum(vals))

    wb.save("欠席.xlsx")
    print("生成完了：欠席.xlsx")

if __name__ == "__main__":
    export_excel()