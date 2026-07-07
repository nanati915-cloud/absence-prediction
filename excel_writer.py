from openpyxl import load_workbook
from openpyxl.drawing.image import Image

class ExcelWriter:
    """
    【概要】分析結果をExcel帳票（テンプレート）へ出力するための専用汎用クラス
    データの集計ロジックとExcelの操作・描画ロジックを完全に分離（カプセル化）することで、
    将来的にデザインやレイアウト変更があっても、このクラスを呼び出すメイン側のコードを一切汚さずに修正できる設計にしています。
    """

    def __init__(self, template_path, output_path):
        self.template_path = template_path
        self.output_path = output_path
        """
        初期化メソッド。
        あらかじめ用意したExcelの「報告書ひな形（テンプレート）」を読み込み、
        出力先パス（output_path）をクラス内に保持します。
        """
        self.workbook = load_workbook(template_path)

    # ======================================================
    # ■ 1. 単一セルへのテキスト書き込み
    # ======================================================
    def write_cell(self, ws, cell_coord, text):
        """
        ws: Worksheetオブジェクト（report_generatorから渡されたもの）
        cell_coord: セル番地（文字列）
        text: 書き込む内容
        """
        # 座標からセルを取得
        cell = ws[cell_coord]
        
        # 結合セルかどうかを判定し、結合なら左上のセルに書き込む
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # 結合範囲の左上のセル番地を取得して再取得
                cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
                
        # 値を書き込む
        cell.value = text

    # ======================================================
    # ■ 2. 解析グラフ画像の一括挿入（自動縦横比維持機能）
    # ======================================================
    def insert_image(self, sheet, image_path, cell="B4", width=520, height=None):
        """
        matplotlib等で出力した分析グラフの画像ファイルを、指定セルの位置へ自動で貼り付けます。
        
        【工夫点】
        グラフの見た目が潰れる（歪む）のを防ぐため、高さ(height)が指定されていない(Noneの)場合、
        元画像の「アスペクト比（縦横比）」をプログラムが自動で逆算します。
        指定された横幅(width)に合わせて縦幅を整数値で動的にリサイズすることで、
        常に「人間の目で見やすい高精度なグラフレイアウト」を担保したままExcelに埋め込むことが可能です。
        """
        ws = self.workbook[sheet]
        img = Image(image_path)

        if height is None:
            # 元画像の縦横比から、横幅に連動した適切な高さを自動計算する（過学習グラフ等の歪み防止）
            aspect_ratio = img.height / img.width
            img.width = width
            img.height = int(width * aspect_ratio)
        else:
            # 明示的に縦横サイズがミリ単位などで指定された場合はそのサイズをそのまま適用
            img.width = width
            img.height = height

        ws.add_image(img, cell)

    # ======================================================
    # ■ 3. データテーブル（二次元配列）の高速マッピング書き込み
    # ======================================================
    def write_table(self, sheet, start_row, start_col, table):
        """
        Pythonの二次元配列（リストのリスト）として渡された統計データやAI予測結果データを、
        Excelのセルへ行・列のインデックスをループ計算しながら自動で順番にマッピング展開します。
        
        【利点】
        「何行目の何列目からデータを展開するか（start_row, start_col）」を指定するだけで、
        大量の数値データをワンショットで表の中に綺麗に配置することができます。
        """
        ws = self.workbook[sheet]
        for i, row in enumerate(table):
            for j, value in enumerate(row):
                ws.cell(
                    row=start_row + i,
                    column=start_col + j,
                    value=value
                )

    # ======================================================
    # ■ 4. 編集内容の確定と最終ファイル保存
    # ======================================================
    def save(self):
        """
        メモリ上で編集・作成したすべてのデータ・画像・テキストを、
        事前にconfigで定義した出力パス（日付付きの成果物Excelファイル）として安全に物理保存します。
        """
        self.workbook.save(self.output_path)